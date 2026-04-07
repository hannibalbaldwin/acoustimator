"""Core Excel buildup parser for Acoustimator.

Reads Commercial Acoustics Excel buildup files, sends structured cell data to
Claude API for intelligent field extraction, and returns validated Pydantic models
matching the database schema.

Handles all four buildup format types:
  - Format A: Simple single-scope (vertical key-value pairs)
  - Format B: Multi-scope with tags (tabular, ACT-1/AWP-1 style)
  - Format C: Complex multi-building (multiple sheets, building/floor breakdowns)
  - Format D: Tabular takeoff (proper column headers, scrap rate column)

Usage:
    from src.extraction.excel_parser import extract_buildup, extract_all_buildups

    result = await extract_buildup(Path("path/to/buildup.xlsx"), "Client Folder")
    results = await extract_all_buildups(Path("/path/to/+ITBs"), concurrency=5)
"""

from __future__ import annotations

import asyncio
import json
import logging
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import anthropic
import openpyxl
from pydantic import BaseModel, field_validator

from src.config import settings

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MODEL = "claude-sonnet-4-6"
MAX_TOKENS = 8192

# Files to skip during batch extraction
SKIP_FILENAME_PATTERNS = re.compile(
    r"(?i)(vendor|quote|template|T-004|~\$|\.tmp)",
)

# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------


class ExtractedScope(BaseModel):
    """A single scope/line item extracted from a buildup."""

    tag: str | None = None  # e.g., "ACT-1", "AWP-1"
    scope_type: str  # ACT, AWP, AP, Baffles, FW, SM, WW, RPG, Other
    product_name: str | None = None  # e.g., "Dune on Suprafine"
    square_footage: Decimal | None = None
    cost_per_sf: Decimal | None = None
    material_cost: Decimal | None = None
    markup_pct: Decimal | None = None  # Stored as decimal: 0.35 = 35%
    material_price: Decimal | None = None
    man_days: Decimal | None = None
    daily_labor_rate: Decimal | None = None
    labor_base_rate: Decimal | None = None
    labor_hours_per_day: Decimal | None = None
    labor_multiplier: Decimal | None = None
    labor_price: Decimal | None = None
    sales_tax_pct: Decimal | None = None
    sales_tax: Decimal | None = None
    county_surtax_rate: Decimal | None = None
    county_surtax_cap: Decimal | None = None
    scrap_rate: Decimal | None = None
    total: Decimal | None = None
    notes: str | None = None
    drawing_references: list[str] = []
    source_sheet: str | None = None

    @field_validator("*", mode="before")
    @classmethod
    def coerce_decimals(cls, v: Any, info: Any) -> Any:
        """Coerce numeric strings and floats to Decimal for Decimal fields."""
        if info.field_name in {
            "tag",
            "scope_type",
            "product_name",
            "notes",
            "drawing_references",
            "source_sheet",
        }:
            return v
        if v is None:
            return v
        if isinstance(v, (int, float, str)):
            try:
                return Decimal(str(v))
            except (InvalidOperation, ValueError):
                return v
        return v


class ExtractedProject(BaseModel):
    """Full extraction result from a buildup file."""

    project_name: str
    folder_name: str
    source_file: str
    format_type: str  # A, B, C, D
    project_address: str | None = None
    gc_name: str | None = None
    gc_contact: str | None = None
    estimator: str | None = None
    sales_rep: str | None = None
    bid_due_date: str | None = None
    scopes: list[ExtractedScope]
    additional_costs: list[dict[str, Any]] = []
    grand_total: Decimal | None = None
    extraction_confidence: float = 0.0
    raw_notes: str | None = None

    @field_validator("grand_total", mode="before")
    @classmethod
    def coerce_grand_total(cls, v: Any) -> Any:
        if v is None:
            return v
        try:
            return Decimal(str(v))
        except (InvalidOperation, ValueError):
            return v


class ExtractionResult(BaseModel):
    """Result of an extraction attempt."""

    success: bool
    project: ExtractedProject | None = None
    error: str | None = None
    tokens_used: int = 0
    model_used: str = MODEL


# ---------------------------------------------------------------------------
# Excel reading
# ---------------------------------------------------------------------------


def format_cell_contents(value: Any) -> str:
    """Return a concise string representation of a cell value.

    Handles None, floats (cleaning IEEE-754 noise), booleans, and strings.
    Strips leading/trailing whitespace from string values.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        # Avoid floating-point noise: 0.30000000000000004 -> "0.3"
        if value == int(value) and abs(value) < 1e12:
            return str(int(value))
        # Round to 6 significant figures to clean up IEEE-754 artifacts
        return f"{value:.6g}"
    result = str(value)
    return result.strip() if isinstance(value, str) else result


# Keep internal alias for backward compat
_cell_repr = format_cell_contents


def read_workbook_cells(file_path: Path) -> tuple[str, str]:
    """Read all cells from an Excel workbook, returning both values and formulas.

    Args:
        file_path: Path to the .xlsx file.

    Returns:
        Tuple of (values_text, formulas_text) where each is a structured text
        representation of all non-empty cells across all sheets.

    Raises:
        openpyxl.utils.exceptions.InvalidFileException: If file is not a valid xlsx.
        FileNotFoundError: If file does not exist.
    """
    # Read calculated values
    wb_data = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    values_lines: list[str] = []

    for sheet_name in wb_data.sheetnames:
        ws = wb_data[sheet_name]
        values_lines.append(f'\nSheet: "{sheet_name}"')
        row_count = 0
        for row in ws.iter_rows():
            cells = []
            for cell in row:
                if cell.value is not None:
                    cells.append(f"{cell.coordinate}: {_cell_repr(cell.value)}")
            if cells:
                values_lines.append("  " + "  |  ".join(cells))
            row_count += 1
            if row_count > 200:
                values_lines.append("  ... (truncated, >200 rows)")
                break
    wb_data.close()

    # Read formulas
    wb_formula = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
    formula_lines: list[str] = []

    for sheet_name in wb_formula.sheetnames:
        ws = wb_formula[sheet_name]
        formula_lines.append(f'\nSheet: "{sheet_name}"')
        has_formulas = False
        row_count = 0
        for row in ws.iter_rows():
            cells = []
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("="):
                    cells.append(f"{cell.coordinate}: {cell.value}")
            if cells:
                formula_lines.append("  " + "  |  ".join(cells))
                has_formulas = True
            row_count += 1
            if row_count > 200:
                break
        if not has_formulas:
            formula_lines.append("  (no formulas)")
    wb_formula.close()

    return "\n".join(values_lines), "\n".join(formula_lines)


# ---------------------------------------------------------------------------
# Claude API prompt
# ---------------------------------------------------------------------------

SYSTEM_PROMPT = """\
You are an expert construction estimating data extractor for Commercial Acoustics, \
a Tampa FL acoustical contractor. You extract structured cost data from Excel buildup \
spreadsheets.

## Buildup Format Types

There are 4 format types. Identify which one applies:

### Format A — Simple Single-Scope
- Vertical key-value pairs in 1-2 columns
- Labels in one column (e.g., "SF", "Cost/SF", "Material Cost"), values in the adjacent column
- One scope per sheet (or a small number of scopes stacked vertically)
- ~30% of all buildups

### Format B — Multi-Scope with Tags
- Multiple scopes in a single sheet, each with a tag like ACT-1, AWP-1, CL-2, etc.
- May have tabular layout or grouped vertical sections per scope
- Summary/grand total row at the bottom
- ~50% of all buildups

### Format C — Complex Multi-Building/Multi-Sheet
- Multiple sheets per workbook (by building, floor, product option, or scope type)
- Building/floor-level breakdowns within sheets
- A summary/total sheet that references other sheets
- ~15% of all buildups

### Format D — Tabular Takeoff
- True tabular format with column headers in row 1
- One line item per row, columns for Category, Name, SF, Cost/SF, Markup, etc.
- May include scrap rate and split material costs (tile vs. grid)
- ~5% of all buildups

## Field Label Variations

The same field may appear with different labels. Map these:

**Square Footage:** SF, Area (SF), Total SF, Vendor SF, Square Footage, Qty, Quantity
**Cost per SF:** Cost/SF, $/SF, Material Cost/SF, Tile Cost/SF, Unit Cost, Cost Each, Cost/Piece
**Material Cost:** Material Cost, Mat Cost, Tile Cost, Grid Cost, Total Cost, Speaker Cost, AVL Cost
**Markup:** Markup, Markup %, Margin, Material Markup, Standard Labor Markup, Standard Travel Markup
**Material Price:** Material Price, Mat Price, Speaker Price, Added Price
**Man-Days:** Man-Days, Install Man-Days, Total Man-Days, MD
**Labor Price:** Labor Price, Labor Cost, Installation Price, Install Price
**Labor Rate:** Labor rate, Base Rate, $/Hr, Hourly Rate
**Hours/Day:** Hours/day, Hrs/Day
**Labor Multiplier:** Multiplier, Burden, Overhead Factor
**Sales Tax:** Sales Tax, Tax, State Sales Tax
**Sales Tax %:** 0.06, 6%, State Sales Tax rate
**County Surtax:** County Surtax, County (Hillsborough), Surtax
**Grand Total:** Total, Grand Total, Project Total
**Scrap Rate:** Scrap, Scrap Rate, Waste Factor, expressed as -5% or 0.10 etc.
**Drawing References:** Page, Sheet, Dwg, Drawing, Floors, Locations

## Scope Type Codes

Map these tags/codes to canonical scope types:
- ACT, LA → ACT (Acoustical Ceiling Tile)
- AWP → AWP (Acoustic Wall Panels)
- AP → AP (Acoustic Panels, custom)
- ACB, CL (when baffle product) → Baffles (Ceiling Baffles)
- FW, SF (when fabric wall) → FW (Fabric Wall)
- SM → SM (Sound Masking)
- WW → WW (WoodWorks)
- RPG → RPG (Specialty Diffusers)
- CL (when ceiling tile/cloud product) → ACT
- Everything else → Other

## Labor Formula

Labor is calculated as: Man-Days × Hours/Day × Base Rate × Multiplier
Common patterns:
- 8 × $45 × 1.45 = $522/day (most common)
- 8 × $45 × 1.50 = $540/day
- 8 × $45 × 1.35 = $486/day
- 8 × $40 × 1.35 = $432/day (ACT sheet pattern)
- 10 × $46 × variable (Sound Masking pattern)

Extract the individual components (hours, base rate, multiplier) when visible in formulas.

## Sales Tax

Florida standard is 6%. Some projects add a county discretionary surtax:
- Hillsborough County: 6% + 1.5% capped at $5,000
- Formula: `0.06 × MaterialPrice + MIN(0.015 × MaterialPrice, 5000)`
- Some entities are tax-exempt (schools, non-profits)

Always extract the actual tax formula/rate used, not assume 6%.

## Additional Costs (Non-Standard Line Items)

Look for these outside the core material/labor/tax formula:
- Lift Rental, Scissor Lift
- Travel: Per Diem, Flights, Hotels, Car/Uber, Truck Rental
- Equipment Rental (Scissor Boom, Compressor, Table Saw)
- Consumables (adhesive, caulk, fasteners)
- P&P Bond (Payment & Performance bond, typically 3%)
- Site Visit
- Punch List / Go-Back
- Setup / Unload / Mobilization
- Commission / Tune / Balance (Sound Masking)
- Axiom / trim / edge molding add-ons
- Freight / Shipping
- S&H (Shipping & Handling)

## Instructions

1. Identify the format type (A, B, C, or D).
2. Extract ALL scopes with ALL available fields. Do not skip any scope.
3. For each scope, extract the individual formula components, not just final values.
4. Identify non-standard/additional costs and categorize them.
5. Extract project metadata (name, address, GC, contact, due date) if present.
6. Assign a confidence score (0.0-1.0) based on:
   - How clearly the data maps to known fields (higher = clearer)
   - Whether calculated values are internally consistent
   - Whether any fields are ambiguous or missing

## Output Format

Return valid JSON matching this exact structure (no markdown, no code blocks):

{
  "project_name": "string",
  "format_type": "A|B|C|D",
  "project_address": "string or null",
  "gc_name": "string or null",
  "gc_contact": "string or null",
  "estimator": "string or null",
  "sales_rep": "string or null",
  "bid_due_date": "string or null",
  "scopes": [
    {
      "tag": "string or null",
      "scope_type": "ACT|AWP|AP|Baffles|FW|SM|WW|RPG|Other",
      "product_name": "string or null",
      "square_footage": number_or_null,
      "cost_per_sf": number_or_null,
      "material_cost": number_or_null,
      "markup_pct": number_or_null,
      "material_price": number_or_null,
      "man_days": number_or_null,
      "daily_labor_rate": number_or_null,
      "labor_base_rate": number_or_null,
      "labor_hours_per_day": number_or_null,
      "labor_multiplier": number_or_null,
      "labor_price": number_or_null,
      "sales_tax_pct": number_or_null,
      "sales_tax": number_or_null,
      "county_surtax_rate": number_or_null,
      "county_surtax_cap": number_or_null,
      "scrap_rate": number_or_null,
      "total": number_or_null,
      "notes": "string or null",
      "drawing_references": ["string"],
      "source_sheet": "string or null"
    }
  ],
  "additional_costs": [
    {
      "cost_type": "lift_rental|travel_per_diem|travel_flights|travel_hotels|equipment|consumables|bond|site_visit|punch_list|setup_unload|commission|other",
      "description": "string",
      "amount": number
    }
  ],
  "grand_total": number_or_null,
  "extraction_confidence": 0.0_to_1.0,
  "raw_notes": "string or null"
}

IMPORTANT:
- markup_pct must be a decimal (0.35 for 35%), NOT a percentage
- sales_tax_pct must be a decimal (0.06 for 6%)
- county_surtax_rate must be a decimal (0.015 for 1.5%)
- scrap_rate must be a decimal (0.10 for 10%)
- All monetary values should be numbers, not formatted strings
- If a Sound Masking scope has a detailed component breakdown, include the totals, \
not individual components
- Include the source_sheet name for each scope when the workbook has multiple sheets\
"""


def _build_user_prompt(
    file_path: Path,
    folder_name: str,
    values_text: str,
    formulas_text: str,
) -> str:
    """Build the user message for the Claude API call."""
    return f"""\
Extract structured cost data from this Commercial Acoustics Excel buildup.

**File:** {file_path.name}
**Folder:** {folder_name}

## Cell Values (calculated/final values)

{values_text}

## Formulas (Excel formulas for cells that have them)

{formulas_text}

Extract all scopes and additional costs. Return only valid JSON.\
"""


# ---------------------------------------------------------------------------
# Core extraction
# ---------------------------------------------------------------------------


def _parse_json_from_response(text: str) -> dict[str, Any]:
    """Extract JSON from Claude's response, handling markdown code blocks.

    Args:
        text: The raw response text from Claude.

    Returns:
        Parsed JSON as a dictionary.

    Raises:
        json.JSONDecodeError: If no valid JSON can be found.
    """
    # Try direct parse first
    text = text.strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # Try extracting from markdown code block
    match = re.search(r"```(?:json)?\s*\n?(.*?)\n?```", text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(1).strip())
        except json.JSONDecodeError:
            pass

    # Try finding the first { ... } block
    brace_start = text.find("{")
    brace_end = text.rfind("}")
    if brace_start != -1 and brace_end != -1 and brace_end > brace_start:
        try:
            return json.loads(text[brace_start : brace_end + 1])
        except json.JSONDecodeError:
            pass

    raise json.JSONDecodeError("No valid JSON found in response", text, 0)


async def extract_buildup(file_path: Path, folder_name: str) -> ExtractionResult:
    """Extract structured data from an Excel buildup file.

    Opens the Excel file with openpyxl (both data_only=True for calculated values
    and data_only=False for formulas), formats cell contents as structured text,
    sends to Claude API for intelligent extraction, and returns validated Pydantic
    models.

    Args:
        file_path: Path to the .xlsx buildup file.
        folder_name: Name of the parent folder (used as fallback project name).

    Returns:
        ExtractionResult with success=True and populated project on success,
        or success=False with error message on failure.
    """
    logger.info("Extracting buildup: %s", file_path)

    # --- Step 1: Read the Excel file ---
    try:
        values_text, formulas_text = read_workbook_cells(file_path)
    except FileNotFoundError:
        return ExtractionResult(
            success=False,
            error=f"File not found: {file_path}",
        )
    except Exception as e:
        logger.exception("Failed to read Excel file: %s", file_path)
        return ExtractionResult(
            success=False,
            error=f"Excel read error: {type(e).__name__}: {e}",
        )

    if not values_text.strip():
        return ExtractionResult(
            success=False,
            error="Workbook contains no data",
        )

    # --- Step 2: Call Claude API (with exponential backoff on rate limit) ---
    user_message = _build_user_prompt(file_path, folder_name, values_text, formulas_text)

    client = anthropic.AsyncAnthropic(api_key=settings.anthropic_api_key)
    response = None
    for attempt in range(4):  # up to 3 retries
        try:
            response = await client.messages.create(
                model=MODEL,
                max_tokens=MAX_TOKENS,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": user_message}],
            )
            break  # success
        except anthropic.RateLimitError:
            if attempt == 3:
                return ExtractionResult(
                    success=False,
                    error="Anthropic API rate limit exceeded after 3 retries",
                )
            wait = 30 * (2**attempt)  # 30s, 60s, 120s
            logger.warning("Rate limited on %s — waiting %ds (attempt %d/3)", file_path.name, wait, attempt + 1)
            await asyncio.sleep(wait)
        except anthropic.AuthenticationError:
            return ExtractionResult(
                success=False,
                error="Anthropic API authentication failed — check ANTHROPIC_API_KEY",
            )
        except anthropic.APIConnectionError as e:
            return ExtractionResult(
                success=False,
                error=f"Anthropic API connection error: {e}",
            )
        except anthropic.APIError as e:
            logger.exception("Anthropic API error during extraction of %s", file_path)
            return ExtractionResult(
                success=False,
                error=f"Anthropic API error: {type(e).__name__}: {e}",
            )

    if response is None:
        return ExtractionResult(success=False, error="No response received after retries")

    # Calculate token usage
    tokens_used = (response.usage.input_tokens or 0) + (response.usage.output_tokens or 0)

    # Extract text from response
    response_text = ""
    for block in response.content:
        if hasattr(block, "text"):
            response_text += block.text

    if not response_text.strip():
        return ExtractionResult(
            success=False,
            error="Claude returned an empty response",
            tokens_used=tokens_used,
            model_used=MODEL,
        )

    # --- Step 3: Parse JSON response ---
    try:
        data = _parse_json_from_response(response_text)
    except json.JSONDecodeError as e:
        logger.error(
            "Failed to parse JSON from Claude response for %s: %s\nResponse: %s",
            file_path,
            e,
            response_text[:500],
        )
        return ExtractionResult(
            success=False,
            error=f"JSON parse error: {e}",
            tokens_used=tokens_used,
            model_used=MODEL,
        )

    # --- Step 4: Validate into Pydantic models ---
    try:
        scopes = [ExtractedScope(**scope_data) for scope_data in data.get("scopes", [])]

        project = ExtractedProject(
            project_name=data.get("project_name", folder_name),
            folder_name=folder_name,
            source_file=str(file_path),
            format_type=data.get("format_type", "unknown"),
            project_address=data.get("project_address"),
            gc_name=data.get("gc_name"),
            gc_contact=data.get("gc_contact"),
            estimator=data.get("estimator"),
            sales_rep=data.get("sales_rep"),
            bid_due_date=data.get("bid_due_date"),
            scopes=scopes,
            additional_costs=data.get("additional_costs", []),
            grand_total=data.get("grand_total"),
            extraction_confidence=data.get("extraction_confidence", 0.0),
            raw_notes=data.get("raw_notes"),
        )
    except Exception as e:
        logger.exception("Pydantic validation error for %s", file_path)
        return ExtractionResult(
            success=False,
            error=f"Validation error: {type(e).__name__}: {e}",
            tokens_used=tokens_used,
            model_used=MODEL,
        )

    logger.info(
        "Extracted %d scopes from %s (confidence=%.2f, format=%s)",
        len(project.scopes),
        file_path.name,
        project.extraction_confidence,
        project.format_type,
    )

    return ExtractionResult(
        success=True,
        project=project,
        tokens_used=tokens_used,
        model_used=MODEL,
    )


# ---------------------------------------------------------------------------
# Batch extraction
# ---------------------------------------------------------------------------


def find_buildup_files(source_dir: Path) -> list[tuple[Path, str]]:
    """Walk directory tree and find .xlsx files that look like buildups.

    Args:
        source_dir: Root directory to search (e.g., the +ITBs folder).

    Returns:
        List of (file_path, folder_name) tuples. The folder_name is the
        immediate parent directory name, used as a fallback project identifier.
    """
    results: list[tuple[Path, str]] = []

    if not source_dir.is_dir():
        logger.error("Source directory does not exist: %s", source_dir)
        return results

    for xlsx_path in sorted(source_dir.rglob("*.xlsx")):
        filename = xlsx_path.name

        # Skip temp files, vendor quotes, and non-buildup files
        if filename.startswith("~$"):
            continue
        if SKIP_FILENAME_PATTERNS.search(filename):
            logger.debug("Skipping non-buildup file: %s", xlsx_path)
            continue

        # Skip files inside Archive subdirectories (they are superseded)
        parts = xlsx_path.relative_to(source_dir).parts
        if "Archive" in parts or "++Archive" in parts:
            logger.debug("Skipping archived file: %s", xlsx_path)
            continue

        # The folder_name is the top-level project folder
        folder_name = parts[0] if parts else xlsx_path.parent.name

        results.append((xlsx_path, folder_name))

    logger.info("Found %d buildup candidates in %s", len(results), source_dir)
    return results


async def extract_all_buildups(
    source_dir: Path,
    concurrency: int = 2,
    request_delay: float = 15.0,
    files: list[tuple[Path, str]] | None = None,
) -> list[ExtractionResult]:
    """Extract all buildups from the ITBs folder with concurrent processing.

    Walks the directory tree finding .xlsx files, skips files that look like
    vendor quotes or non-buildup files, and processes them with semaphore-limited
    concurrency.

    Args:
        source_dir: Root directory containing project folders (e.g., the +ITBs folder).
        concurrency: Maximum number of concurrent Claude API calls.
        request_delay: Seconds to wait between requests to respect token rate limits.
            Default 15s → ~4 requests/min → ~20K input tokens/min (safe for Tier 1).
        files: Pre-computed list of (file_path, folder_name) tuples. If None, walks
            source_dir automatically.

    Returns:
        List of ExtractionResult objects, one per file processed.
    """
    if files is None:
        files = find_buildup_files(source_dir)

    if not files:
        logger.warning("No buildup files found in %s", source_dir)
        return []

    semaphore = asyncio.Semaphore(concurrency)
    results: list[ExtractionResult] = []
    completed = 0
    total = len(files)

    try:
        from rich.progress import BarColumn, Progress, SpinnerColumn, TaskProgressColumn, TextColumn
    except ImportError:
        Progress = None  # type: ignore[assignment, misc]

    async def _process(file_path: Path, folder_name: str) -> ExtractionResult:
        nonlocal completed
        async with semaphore:
            result = await extract_buildup(file_path, folder_name)
            completed += 1
            status = "OK" if result.success else "FAIL"
            if Progress is None:
                logger.info(
                    "[%d/%d] %s — %s (%d tokens)",
                    completed,
                    total,
                    file_path.name,
                    status,
                    result.tokens_used,
                )
            # Pace requests to stay within token rate limits (Tier 1: 30K tokens/min)
            if request_delay > 0:
                await asyncio.sleep(request_delay)
            return result

    if Progress is not None:
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
        ) as progress:
            task = progress.add_task("Extracting buildups...", total=total)

            async def _process_with_progress(file_path: Path, folder_name: str) -> ExtractionResult:
                result = await _process(file_path, folder_name)
                status = "OK" if result.success else "FAIL"
                progress.update(
                    task,
                    advance=1,
                    description=f"[{status}] {file_path.name}",
                )
                return result

            tasks = [_process_with_progress(fp, fn) for fp, fn in files]
            results = await asyncio.gather(*tasks)
    else:
        tasks = [_process(fp, fn) for fp, fn in files]
        results = await asyncio.gather(*tasks)

    # Summary
    success_count = sum(1 for r in results if r.success)
    fail_count = sum(1 for r in results if not r.success)
    total_tokens = sum(r.tokens_used for r in results)
    total_scopes = sum(len(r.project.scopes) for r in results if r.project)

    logger.info(
        "Batch extraction complete: %d/%d succeeded, %d failed, %d scopes, %d tokens",
        success_count,
        total,
        fail_count,
        total_scopes,
        total_tokens,
    )

    return list(results)


# ---------------------------------------------------------------------------
# Demo / standalone test
# ---------------------------------------------------------------------------


async def demo_extract(file_path: str) -> None:
    """Demo extraction on a single file, printing results.

    Useful for testing and debugging. Prints the full extraction result
    with formatted output.

    Args:
        file_path: Path to an .xlsx buildup file.
    """
    path = Path(file_path)
    folder_name = path.parent.name

    print(f"\n{'=' * 70}")
    print(f"Extracting: {path.name}")
    print(f"Folder: {folder_name}")
    print(f"{'=' * 70}\n")

    result = await extract_buildup(path, folder_name)

    if not result.success:
        print(f"EXTRACTION FAILED: {result.error}")
        return

    project = result.project
    assert project is not None

    print(f"Project: {project.project_name}")
    print(f"Format: {project.format_type}")
    print(f"Confidence: {project.extraction_confidence:.0%}")
    print(f"Tokens used: {result.tokens_used}")

    if project.gc_name:
        print(f"GC: {project.gc_name}")
    if project.project_address:
        print(f"Address: {project.project_address}")

    print(f"\n--- Scopes ({len(project.scopes)}) ---")
    for i, scope in enumerate(project.scopes, 1):
        tag = scope.tag or "(no tag)"
        product = scope.product_name or "(no product)"
        print(f"\n  [{i}] {tag} — {scope.scope_type}: {product}")
        if scope.square_footage:
            print(f"      SF: {scope.square_footage}")
        if scope.cost_per_sf:
            print(f"      Cost/SF: ${scope.cost_per_sf}")
        if scope.material_cost:
            print(f"      Material Cost: ${scope.material_cost}")
        if scope.markup_pct:
            print(f"      Markup: {float(scope.markup_pct) * 100:.0f}%")
        if scope.material_price:
            print(f"      Material Price: ${scope.material_price}")
        if scope.man_days:
            print(f"      Man-Days: {scope.man_days}")
        if scope.daily_labor_rate:
            print(f"      Daily Rate: ${scope.daily_labor_rate}")
        if scope.labor_base_rate:
            print(f"      Base Rate: ${scope.labor_base_rate}/hr")
        if scope.labor_hours_per_day:
            print(f"      Hours/Day: {scope.labor_hours_per_day}")
        if scope.labor_multiplier:
            print(f"      Multiplier: {scope.labor_multiplier}")
        if scope.labor_price:
            print(f"      Labor Price: ${scope.labor_price}")
        if scope.sales_tax:
            print(f"      Sales Tax: ${scope.sales_tax}")
        if scope.sales_tax_pct:
            print(f"      Tax Rate: {float(scope.sales_tax_pct) * 100:.1f}%")
        if scope.county_surtax_rate:
            print(f"      County Surtax: {float(scope.county_surtax_rate) * 100:.1f}%")
        if scope.scrap_rate:
            print(f"      Scrap Rate: {float(scope.scrap_rate) * 100:.0f}%")
        if scope.total:
            print(f"      Total: ${scope.total}")
        if scope.notes:
            print(f"      Notes: {scope.notes}")
        if scope.source_sheet:
            print(f"      Sheet: {scope.source_sheet}")

    if project.additional_costs:
        print(f"\n--- Additional Costs ({len(project.additional_costs)}) ---")
        for cost in project.additional_costs:
            desc = cost.get("description", cost.get("cost_type", "?"))
            amount = cost.get("amount", "?")
            print(f"  {desc}: ${amount}")

    if project.grand_total:
        print(f"\n  GRAND TOTAL: ${project.grand_total}")

    if project.raw_notes:
        print(f"\n--- Notes ---\n  {project.raw_notes}")

    print()


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------


if __name__ == "__main__":
    import sys

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )

    if len(sys.argv) < 2:
        print("Usage: python -m src.extraction.excel_parser <path-to-xlsx>")
        print("       python -m src.extraction.excel_parser --batch <itbs-dir> [concurrency]")
        sys.exit(1)

    if sys.argv[1] == "--batch":
        source_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else settings.data_source_path
        concurrency = int(sys.argv[3]) if len(sys.argv) > 3 else 5
        results = asyncio.run(extract_all_buildups(source_dir, concurrency))
        success = sum(1 for r in results if r.success)
        print(f"\nDone: {success}/{len(results)} succeeded")
    else:
        asyncio.run(demo_extract(sys.argv[1]))
