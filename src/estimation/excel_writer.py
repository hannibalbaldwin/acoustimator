"""Excel export for Acoustimator estimates — Format B buildup style.

Generates a Commercial Acoustics-style "buildup" spreadsheet from a
ProjectEstimate, matching the multi-scope Format B layout used internally.

Usage::

    from pathlib import Path
    from src.estimation.excel_writer import write_estimate_to_excel

    output = write_estimate_to_excel(
        estimate,
        output_path=Path("data/exports/my_project.xlsx"),
        project_name="Hillsborough County Courthouse",
        gc_name="Skanska USA",
    )
    print(f"Wrote estimate to {output}")
"""

from __future__ import annotations

import logging
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill

from src.estimation.models import ProjectEstimate, ScopeEstimate

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

FONT_NAME = "Calibri"
FONT_SIZE_NORMAL = 11
FONT_SIZE_GRAND_TOTAL = 14

FILL_HEADER = PatternFill(fill_type="solid", fgColor="D9D9D9")  # light gray
FILL_SCOPE_TAG = PatternFill(fill_type="solid", fgColor="FFFACD")  # lemon chiffon yellow

FMT_MONEY = '"$"#,##0.00'
FMT_QTY = "#,##0"
FMT_PCT = "0.0%"

COL_WIDTHS = {
    "A": 35,
    "B": 10,
    "C": 8,
    "D": 12,
    "E": 14,
    "F": 30,
}

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _to_float(value: Any) -> float | None:
    """Safely convert a Decimal, float, int, or None to float."""
    if value is None:
        return None
    if isinstance(value, Decimal):
        try:
            return float(value)
        except (InvalidOperation, ValueError):
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _font(bold: bool = False, size: int = FONT_SIZE_NORMAL) -> Font:
    return Font(name=FONT_NAME, size=size, bold=bold)


def _write_row(
    ws: Any,
    row: int,
    values: list[Any],
    bold: bool = False,
    fill: PatternFill | None = None,
    font_size: int = FONT_SIZE_NORMAL,
    number_formats: dict[int, str] | None = None,
) -> None:
    """Write a list of values into columns A–F of the given row.

    Args:
        ws: openpyxl worksheet.
        row: 1-based row number.
        values: Up to 6 values for columns A–F.  Extra values are ignored.
        bold: Whether to apply bold font to all cells in the row.
        fill: Optional background fill for all cells in the row.
        font_size: Font size for all cells.
        number_formats: Mapping of 1-based column index → Excel number format string.
    """
    for col_idx, value in enumerate(values[:6], start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = _font(bold=bold, size=font_size)
        if fill:
            cell.fill = fill
        if number_formats and col_idx in number_formats:
            cell.number_format = number_formats[col_idx]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def write_estimate_to_excel(
    estimate: ProjectEstimate,
    output_path: Path,
    project_name: str = "",
    gc_name: str = "",
    daily_labor_rate: float = 725.0,
    sales_tax_pct: float = 0.075,
    scrap_pct: float = 0.10,
    include_comparables: bool = False,
) -> Path:
    """Write a ProjectEstimate to a Format B Excel buildup.

    Produces a two-sheet workbook:
    - Sheet 1 ("Estimate"): Format B buildup with per-scope material/markup/labor/tax rows.
    - Sheet 2 ("Notes"): All estimate notes and (optionally) comparable projects.

    Args:
        estimate: The ProjectEstimate to export.
        output_path: Destination .xlsx file path.
        project_name: Project name for the header (overrides estimate.source_plan if set).
        gc_name: General Contractor name for the header.
        daily_labor_rate: Loaded daily labor rate in $/day (default $725).
        sales_tax_pct: Florida sales tax fraction (default 0.075 = 7.5%).
        scrap_pct: Material scrap/waste factor (default 0.10 = 10%).
        include_comparables: If True, list comparable projects on the Notes sheet.

    Returns:
        output_path (for chaining).
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws_est = wb.active
    ws_est.title = "Estimate"
    ws_notes = wb.create_sheet("Notes")

    # ------------------------------------------------------------------
    # Derive display project name
    # ------------------------------------------------------------------
    display_name = project_name or estimate.source_plan or ""
    display_gc = gc_name or ""

    # ------------------------------------------------------------------
    # Sheet 1: Estimate
    # ------------------------------------------------------------------
    _build_estimate_sheet(
        ws=ws_est,
        estimate=estimate,
        display_name=display_name,
        display_gc=display_gc,
        daily_labor_rate=daily_labor_rate,
        sales_tax_pct=sales_tax_pct,
        scrap_pct=scrap_pct,
    )

    # ------------------------------------------------------------------
    # Sheet 2: Notes
    # ------------------------------------------------------------------
    _build_notes_sheet(
        ws=ws_notes,
        estimate=estimate,
        include_comparables=include_comparables,
    )

    # ------------------------------------------------------------------
    # Column widths (both sheets inherit same widths for the estimate)
    # ------------------------------------------------------------------
    for col_letter, width in COL_WIDTHS.items():
        ws_est.column_dimensions[col_letter].width = width
    # Notes sheet: just widen column A
    ws_notes.column_dimensions["A"].width = 80

    wb.save(output_path)
    logger.info("Wrote Format B estimate to %s", output_path)
    return output_path


# ---------------------------------------------------------------------------
# Private: estimate sheet builder
# ---------------------------------------------------------------------------


def _build_estimate_sheet(
    ws: Any,
    estimate: ProjectEstimate,
    display_name: str,
    display_gc: str,
    daily_labor_rate: float,
    sales_tax_pct: float,
    scrap_pct: float,
) -> None:
    """Populate the main Estimate worksheet."""

    # ---- Rows 1–3: Header ---------------------------------------------------
    _write_row(ws, 1, ["PROJECT:", display_name], bold=True)
    _write_row(ws, 2, ["DATE:", str(date.today())])
    _write_row(ws, 3, ["GC:", display_gc])
    # Row 4: blank separator (already empty)

    # ---- Row 5: Column headers ---------------------------------------------
    _write_row(
        ws,
        5,
        ["Description", "Qty", "Unit", "Unit Cost", "Total", "Notes"],
        bold=True,
        fill=FILL_HEADER,
    )

    # ---- Scope sections ----------------------------------------------------
    current_row = 6
    scope_subtotal_cells: list[str] = []  # collect E-column subtotal cell addresses

    active_scopes = [s for s in estimate.scope_estimates if s.total is not None]

    if not active_scopes:
        # Graceful empty-state
        ws.cell(row=current_row, column=1, value="No scopes extracted from plan.").font = _font(
            bold=True, size=FONT_SIZE_NORMAL
        )
        current_row += 2
    else:
        for scope in active_scopes:
            current_row = _write_scope_section(
                ws=ws,
                scope=scope,
                start_row=current_row,
                daily_labor_rate=daily_labor_rate,
                sales_tax_pct=sales_tax_pct,
                scrap_pct=scrap_pct,
                scope_subtotal_cells=scope_subtotal_cells,
            )

    # ---- Grand Total --------------------------------------------------------
    current_row = _write_grand_total(
        ws=ws,
        row=current_row,
        scope_subtotal_cells=scope_subtotal_cells,
        estimate=estimate,
    )

    # ---- Additional Costs (if any) -----------------------------------------
    additional_costs = getattr(estimate, "additional_costs", None) or []
    if additional_costs:
        current_row += 1
        _write_row(ws, current_row, ["Additional Costs"], bold=True, fill=FILL_HEADER)
        current_row += 1
        addl_cells: list[str] = []
        for cost in additional_costs:
            amount = _to_float(getattr(cost, "amount", None))
            ws.cell(row=current_row, column=1, value=f"  {cost.description}").font = _font()
            if amount is not None:
                e_cell = ws.cell(row=current_row, column=5, value=amount)
                e_cell.font = _font()
                e_cell.number_format = FMT_MONEY
                addl_cells.append(e_cell.coordinate)
            current_row += 1

        if addl_cells:
            current_row += 1
            addl_total_cell = ws.cell(
                row=current_row,
                column=5,
                value=f"=SUM({','.join(addl_cells)})",
            )
            addl_total_cell.font = _font(bold=True)
            addl_total_cell.number_format = FMT_MONEY
            ws.cell(row=current_row, column=1, value="  Additional Costs Total").font = _font(
                bold=True
            )
            current_row += 1


def _write_scope_section(
    ws: Any,
    scope: ScopeEstimate,
    start_row: int,
    daily_labor_rate: float,
    sales_tax_pct: float,
    scrap_pct: float,
    scope_subtotal_cells: list[str],
) -> int:
    """Write one scope's rows (header, material, markup, labor, tax, subtotal).

    Returns the next available row number after writing this scope + blank separator.
    """
    row = start_row

    # Resolve numeric values from Decimal fields
    area_sf = _to_float(scope.area_sf)
    cost_per_sf = _to_float(scope.predicted_cost_per_sf)
    markup_pct = _to_float(scope.predicted_markup_pct)
    man_days = _to_float(scope.predicted_man_days)
    material_cost = _to_float(scope.material_cost)

    # Fall back to deriving material_cost if not pre-computed
    if material_cost is None and area_sf is not None and cost_per_sf is not None:
        material_cost = area_sf * cost_per_sf * (1.0 + scrap_pct)

    if markup_pct is None:
        markup_pct = 0.33  # sensible default

    # ---- Scope header row --------------------------------------------------
    tag_label = scope.scope_tag or scope.scope_type
    description = f"{tag_label} - {scope.product_hint}" if scope.product_hint else tag_label
    _write_row(ws, row, [description], bold=True, fill=FILL_SCOPE_TAG)
    row += 1

    # Track E-column row numbers for SUM in subtotal
    scope_row_cells: list[str] = []

    # ---- Material row -------------------------------------------------------
    mat_row = row
    if area_sf is not None and cost_per_sf is not None:
        mat_ext_formula = f"=B{mat_row}*D{mat_row}"
        _write_row(
            ws,
            mat_row,
            [
                "  Materials",
                area_sf * (1.0 + scrap_pct),  # qty includes scrap
                "SF",
                cost_per_sf,
                mat_ext_formula,
                scope.product_hint or "",
            ],
            number_formats={2: FMT_QTY, 4: FMT_MONEY, 5: FMT_MONEY},
        )
    elif material_cost is not None:
        # We have material_cost directly but not cost/SF — write as lump sum
        mat_ext_formula = f"=E{mat_row}"
        _write_row(
            ws,
            mat_row,
            ["  Materials", None, "", None, material_cost, scope.product_hint or ""],
            number_formats={5: FMT_MONEY},
        )
    else:
        mat_ext_formula = f"=E{mat_row}"
        _write_row(ws, mat_row, ["  Materials", None, "", None, None, scope.product_hint or ""])

    scope_row_cells.append(f"E{mat_row}")
    row += 1

    # ---- Markup row ---------------------------------------------------------
    mkp_row = row
    markup_formula = f"=E{mat_row}*{markup_pct}"
    _write_row(
        ws,
        mkp_row,
        ["  Markup", markup_pct, "%", "", markup_formula, ""],
        number_formats={2: FMT_PCT, 5: FMT_MONEY},
    )
    scope_row_cells.append(f"E{mkp_row}")
    row += 1

    # ---- Labor row ----------------------------------------------------------
    lbr_row = row
    if man_days is not None:
        lbr_ext_formula = f"=B{lbr_row}*D{lbr_row}"
        _write_row(
            ws,
            lbr_row,
            ["  Labor", man_days, "days", daily_labor_rate, lbr_ext_formula, ""],
            number_formats={4: FMT_MONEY, 5: FMT_MONEY},
        )
    else:
        lbr_cost = _to_float(scope.labor_cost)
        _write_row(
            ws,
            lbr_row,
            ["  Labor", None, "days", daily_labor_rate, lbr_cost, ""],
            number_formats={4: FMT_MONEY, 5: FMT_MONEY},
        )
    scope_row_cells.append(f"E{lbr_row}")
    row += 1

    # ---- Tax row ------------------------------------------------------------
    tax_row = row
    tax_formula = f"=(E{mat_row}+E{mkp_row})*{sales_tax_pct}"
    _write_row(
        ws,
        tax_row,
        [f"  Sales Tax ({sales_tax_pct:.1%})", sales_tax_pct, "%", "", tax_formula, ""],
        number_formats={2: FMT_PCT, 5: FMT_MONEY},
    )
    scope_row_cells.append(f"E{tax_row}")
    row += 1

    # ---- Scope subtotal row ------------------------------------------------
    sub_row = row
    subtotal_formula = f"=SUM({','.join(scope_row_cells)})"
    _write_row(
        ws,
        sub_row,
        ["  SCOPE TOTAL", "", "", "", subtotal_formula, ""],
        bold=True,
        number_formats={5: FMT_MONEY},
    )
    scope_subtotal_cells.append(f"E{sub_row}")
    row += 1

    # ---- Blank separator ---------------------------------------------------
    row += 1

    return row


def _write_grand_total(
    ws: Any,
    row: int,
    scope_subtotal_cells: list[str],
    estimate: ProjectEstimate,
) -> int:
    """Write the grand total row and optional confidence warning.

    Returns the next available row.
    """
    if scope_subtotal_cells:
        grand_total_formula = f"=SUM({','.join(scope_subtotal_cells)})"
    else:
        grand_total_formula = 0

    gt_cell = ws.cell(row=row, column=5, value=grand_total_formula)
    gt_cell.font = _font(bold=True, size=FONT_SIZE_GRAND_TOTAL)
    gt_cell.number_format = FMT_MONEY

    label_cell = ws.cell(row=row, column=1, value="GRAND TOTAL")
    label_cell.font = _font(bold=True, size=FONT_SIZE_GRAND_TOTAL)
    row += 1

    # Low-confidence note
    confidences = [s.confidence for s in estimate.scope_estimates if s.confidence is not None]
    if confidences:
        avg_conf = sum(confidences) / len(confidences)
        if avg_conf < 0.6:
            row += 1
            note_cell = ws.cell(
                row=row,
                column=1,
                value=(
                    f"NOTE: Average prediction confidence is {avg_conf:.0%} — "
                    "review estimates carefully before quoting."
                ),
            )
            note_cell.font = Font(name=FONT_NAME, size=FONT_SIZE_NORMAL, bold=False, color="FF0000")
            row += 1

    return row


# ---------------------------------------------------------------------------
# Private: notes sheet builder
# ---------------------------------------------------------------------------


def _build_notes_sheet(
    ws: Any,
    estimate: ProjectEstimate,
    include_comparables: bool,
) -> None:
    """Populate the Notes worksheet with warnings and optional comparables."""
    ws.cell(row=1, column=1, value="Estimate Notes").font = _font(
        bold=True, size=FONT_SIZE_GRAND_TOTAL
    )

    row = 3

    if estimate.notes:
        ws.cell(row=row, column=1, value="Warnings / Flags").font = _font(bold=True)
        row += 1
        for note in estimate.notes:
            ws.cell(row=row, column=1, value=f"  • {note}").font = _font()
            row += 1
        row += 1
    else:
        ws.cell(row=row, column=1, value="  (No warnings)").font = _font()
        row += 2

    # Per-scope confidence summary
    ws.cell(row=row, column=1, value="Scope Confidence Summary").font = _font(bold=True)
    row += 1
    for scope in estimate.scope_estimates:
        tag = scope.scope_tag or scope.scope_type
        conf = scope.confidence
        model = scope.model_used or "—"
        conf_str = f"{conf:.0%}" if conf is not None else "N/A"
        ws.cell(
            row=row,
            column=1,
            value=f"  {tag}: confidence={conf_str}, model={model}",
        ).font = _font()
        row += 1

    # Comparables
    if include_comparables:
        row += 1
        ws.cell(row=row, column=1, value="Comparable Projects").font = _font(bold=True)
        row += 1
        all_comps: list[str] = []
        for scope in estimate.scope_estimates:
            comps = getattr(scope, "comparable_projects", [])
            all_comps.extend(comps)
        if all_comps:
            for comp in all_comps:
                ws.cell(row=row, column=1, value=f"  • {comp}").font = _font()
                row += 1
        else:
            ws.cell(row=row, column=1, value="  (none available)").font = _font()
            row += 1
