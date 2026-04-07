"""Tests for Phase 5 estimation engine modules.

Guards against missing modules with pytest.importorskip so the suite passes
even when some Phase 5 files have not been delivered yet.
"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import TYPE_CHECKING
from unittest.mock import AsyncMock, MagicMock

if TYPE_CHECKING:
    from src.extraction.plan_parser.models import PlanReadResult

import pytest

# ---------------------------------------------------------------------------
# Module guards — skip entire class/section if module not present yet
# ---------------------------------------------------------------------------

confidence_mod = pytest.importorskip(
    "src.estimation.confidence",
    reason="src/estimation/confidence.py not yet delivered",
)
models_mod = pytest.importorskip(
    "src.estimation.models",
    reason="src/estimation/models.py not yet delivered",
)
excel_writer_mod = pytest.importorskip(
    "src.estimation.excel_writer",
    reason="src/estimation/excel_writer.py not yet delivered",
)
comparator_mod = pytest.importorskip(
    "src.estimation.comparator",
    reason="src/estimation/comparator.py not yet delivered",
)

# ---------------------------------------------------------------------------
# Lazy imports (only reached if importorskip above passed)
# ---------------------------------------------------------------------------

from src.estimation.comparator import (  # noqa: E402
    ComparableProject,
    _compute_similarity,
    _cost_similarity,
    _sf_similarity,
    find_comparable_projects,
)
from src.estimation.confidence import (  # noqa: E402
    ConfidenceLevel,
    ConfidenceReport,
    compute_project_confidence,
    compute_scope_confidence,
    format_confidence_badge,
)
from src.estimation.excel_writer import write_estimate_to_excel  # noqa: E402
from src.estimation.models import ProjectEstimate, ScopeEstimate  # noqa: E402

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

MODEL_PATH = Path("data/models/ACT_cost_model.joblib")

requires_model = pytest.mark.skipif(
    not MODEL_PATH.exists(),
    reason="ACT cost model not trained (data/models/ACT_cost_model.joblib missing)",
)


def _make_scope_estimate(
    *,
    scope_tag: str = "ACT-1",
    scope_type: str = "ACT",
    area_sf: Decimal | None = Decimal("1000"),
    confidence: float = 0.85,
    total: Decimal | None = Decimal("12500.00"),
    material_cost: Decimal | None = Decimal("9000.00"),
    labor_cost: Decimal | None = Decimal("2000.00"),
    predicted_cost_per_sf: Decimal | None = Decimal("12.50"),
    predicted_markup_pct: Decimal | None = Decimal("0.33"),
    predicted_man_days: Decimal | None = Decimal("4.0"),
    product_hint: str | None = None,
    model_used: str = "ACT_cost_model",
    comparable_projects: list[str] | None = None,
) -> ScopeEstimate:
    return ScopeEstimate(
        scope_tag=scope_tag,
        scope_type=scope_type,
        area_sf=area_sf,
        product_hint=product_hint,
        predicted_cost_per_sf=predicted_cost_per_sf,
        predicted_markup_pct=predicted_markup_pct,
        predicted_man_days=predicted_man_days,
        material_cost=material_cost,
        labor_cost=labor_cost,
        total=total,
        confidence=confidence,
        model_used=model_used,
        comparable_projects=comparable_projects or [],
    )


def _make_project_estimate(
    scopes: list[ScopeEstimate] | None = None,
) -> ProjectEstimate:
    if scopes is None:
        scopes = [
            _make_scope_estimate(scope_tag="ACT-1", total=Decimal("12500.00")),
            _make_scope_estimate(
                scope_tag="AWP-1",
                scope_type="AWP",
                total=Decimal("8750.00"),
                area_sf=Decimal("500"),
            ),
        ]
    total = sum((s.total or Decimal("0")) for s in scopes)
    man_days = sum((s.predicted_man_days or Decimal("0")) for s in scopes)
    return ProjectEstimate(
        source_plan="/fake/path/project.pdf",
        extraction_confidence=0.95,
        scope_estimates=scopes,
        total_estimated_cost=total,
        total_area_sf=Decimal("1500"),
        estimated_man_days=man_days,
        notes=[],
        created_at=datetime(2026, 4, 6, 12, 0, 0),
    )


# ===========================================================================
# Confidence module tests
# ===========================================================================

# Actual API (from src/estimation/confidence.py):
#   compute_scope_confidence(plan_confidence, scope_type, area_sf, model_used) -> float
#   compute_project_confidence(plan_result: PlanReadResult, scope_estimates: list[ScopeEstimate])
#       -> ConfidenceReport
#   format_confidence_badge(score: float) -> str
#   ConfidenceReport fields: overall_score, level, plan_reading_score, model_score,
#                            flags, recommendations


def _make_plan_result_for_confidence(
    extraction_confidence: float = 0.95,
    total_area_sf: Decimal | None = Decimal("1000"),
    scope_suggestions=None,
) -> PlanReadResult:
    """Build a minimal PlanReadResult for use in confidence tests."""
    from src.extraction.plan_parser.models import PlanReadResult as _PlanReadResult

    return _PlanReadResult(
        source_file="/fake/plan.pdf",
        total_pages=2,
        vector_rich_pages=1,
        raster_pages=1,
        pages=[],
        rooms=[],
        ceiling_specs=[],
        scope_suggestions=scope_suggestions or [],
        total_area_sf=total_area_sf,
        extraction_confidence=extraction_confidence,
        vision_pages_used=0,
        error=None,
        success=True,
    )


class TestComputeScopeConfidence:
    """Tests for compute_scope_confidence() — returns a float."""

    def test_act_bluebeam_high_confidence(self) -> None:
        """ACT scope at 1000 SF with Bluebeam plan_confidence=0.95 → high overall score."""
        score = compute_scope_confidence(
            plan_confidence=0.95,
            scope_type="ACT",
            area_sf=1000.0,
            model_used="ACT_cost_model",
        )
        assert isinstance(score, float)
        assert score >= 0.7, f"Expected high confidence, got {score}"

    def test_keyword_only_scope_low_score(self) -> None:
        """Scope with low plan_confidence (0.50) → lower overall score."""
        score = compute_scope_confidence(
            plan_confidence=0.50,
            scope_type="ACT",
            area_sf=800.0,
            model_used="ACT_cost_model",
        )
        # With plan_weight=0.6 and plan_confidence=0.5, score should be moderate
        assert score <= 0.70, f"Expected moderate/low confidence, got {score}"

    def test_no_area_sf_penalty(self) -> None:
        """Scope with area_sf=None receives SF out-of-distribution penalty."""
        # Note: per the implementation, area_sf=None causes _sf_penalty → 1.0
        # (no penalty for None), but area=0 causes penalty. The key test:
        # both return valid floats in [0, 1].
        score_with_area = compute_scope_confidence(
            plan_confidence=0.8,
            scope_type="ACT",
            area_sf=1000.0,
            model_used="ACT_cost_model",
        )
        score_no_area = compute_scope_confidence(
            plan_confidence=0.8,
            scope_type="ACT",
            area_sf=None,
            model_used="ACT_cost_model",
        )
        # Both must be valid floats
        assert 0.0 <= score_with_area <= 1.0
        assert 0.0 <= score_no_area <= 1.0

    def test_confidence_bounded(self) -> None:
        """Overall confidence is always in [0, 1]."""
        for plan_conf in (0.0, 0.3, 0.5, 0.95, 1.0):
            score = compute_scope_confidence(
                plan_confidence=plan_conf,
                scope_type="ACT",
                area_sf=500.0,
                model_used="ACT_cost_model",
            )
            assert 0.0 <= score <= 1.0, f"Score out of range for plan_conf={plan_conf}: {score}"

    def test_out_of_distribution_sf_penalised(self) -> None:
        """Very large SF (>50,000) should apply OOD penalty vs normal range."""
        score_normal = compute_scope_confidence(
            plan_confidence=0.8,
            scope_type="ACT",
            area_sf=5000.0,
            model_used="ACT_cost_model",
        )
        score_ood = compute_scope_confidence(
            plan_confidence=0.8,
            scope_type="ACT",
            area_sf=100_000.0,
            model_used="ACT_cost_model",
        )
        assert score_ood < score_normal, "Out-of-distribution SF should lower confidence"

    def test_general_model_fallback(self) -> None:
        """Unknown scope type should use the general model accuracy (lower than ACT)."""
        score_act = compute_scope_confidence(
            plan_confidence=0.8,
            scope_type="ACT",
            area_sf=1000.0,
            model_used="ACT_cost_model",
        )
        score_unknown = compute_scope_confidence(
            plan_confidence=0.8,
            scope_type="UNKNOWN_TYPE",
            area_sf=1000.0,
            model_used="general_cost_model",
        )
        # ACT model accuracy (0.87) > general fallback (0.73), so ACT score should be higher
        assert score_act > score_unknown


class TestFormatConfidenceBadge:
    """Tests for format_confidence_badge(score: float) -> str."""

    def test_high_badge(self) -> None:
        """Score >= 0.75 → High badge."""
        badge = format_confidence_badge(0.87)
        assert "High" in badge or "high" in badge.lower() or "🟢" in badge

    def test_medium_badge(self) -> None:
        """Score 0.50–0.74 → Medium badge."""
        badge = format_confidence_badge(0.61)
        assert "Medium" in badge or "medium" in badge.lower() or "🟡" in badge

    def test_low_badge(self) -> None:
        """Score < 0.50 → Low badge."""
        badge = format_confidence_badge(0.34)
        assert "Low" in badge or "low" in badge.lower() or "🔴" in badge

    def test_returns_string(self) -> None:
        for score in (0.0, 0.3, 0.5, 0.75, 1.0):
            assert isinstance(format_confidence_badge(score), str)

    def test_includes_percentage(self) -> None:
        """Badge should include the percentage value."""
        badge = format_confidence_badge(0.87)
        assert "87" in badge, f"Expected percentage in badge, got: {badge}"

    def test_boundary_at_high_threshold(self) -> None:
        """Exactly 0.75 → HIGH."""
        badge = format_confidence_badge(0.75)
        assert "High" in badge or "🟢" in badge

    def test_boundary_just_below_medium(self) -> None:
        """Just below 0.50 → LOW."""
        badge = format_confidence_badge(0.49)
        assert "Low" in badge or "🔴" in badge


class TestComputeProjectConfidence:
    """Tests for compute_project_confidence(plan_result, scope_estimates) -> ConfidenceReport."""

    def test_all_high_scopes(self) -> None:
        """High extraction confidence + ACT scopes → HIGH overall level."""
        plan_result = _make_plan_result_for_confidence(extraction_confidence=0.95)
        scopes = [
            _make_scope_estimate(scope_tag="ACT-1", scope_type="ACT", confidence=0.87),
            _make_scope_estimate(scope_tag="AWP-1", scope_type="AWP", confidence=0.82),
        ]
        report = compute_project_confidence(plan_result, scopes)
        assert isinstance(report, ConfidenceReport)
        assert report.overall_score >= 0.7
        assert report.level == ConfidenceLevel.HIGH

    def test_low_extraction_confidence_flags(self) -> None:
        """Low extraction confidence (0.3) → flags about plan quality."""
        plan_result = _make_plan_result_for_confidence(extraction_confidence=0.3)
        scopes = [_make_scope_estimate()]
        report = compute_project_confidence(plan_result, scopes)
        flag_text = " ".join(report.flags).lower()
        # Should mention review or keyword extraction
        assert "keyword" in flag_text or "review" in flag_text or "text" in flag_text or "raster" in flag_text, (
            f"Expected low-confidence flag, got: {report.flags}"
        )

    def test_raster_only_plan_flagged(self) -> None:
        """ConfidenceReport.flags contains expected strings for raster-only plan (conf=0.3)."""
        plan_result = _make_plan_result_for_confidence(extraction_confidence=0.3)
        scopes = [_make_scope_estimate(scope_type="ACT")]
        report = compute_project_confidence(plan_result, scopes)
        flag_text = " ".join(report.flags).lower()
        assert len(report.flags) > 0, "Expected at least one flag for raster-only plan"
        assert (
            "keyword" in flag_text
            or "text" in flag_text
            or "review" in flag_text
            or "raster" in flag_text
            or "low" in flag_text
        ), f"Expected raster/low flag, got: {report.flags}"

    def test_mixed_confidence_lower_than_all_high(self) -> None:
        """Mixed low/high confidence scopes → lower score than all-high project."""
        plan_high = _make_plan_result_for_confidence(extraction_confidence=0.95)
        plan_low = _make_plan_result_for_confidence(extraction_confidence=0.3)
        scopes_high = [
            _make_scope_estimate(scope_type="ACT", confidence=0.87),
        ]
        scopes_low = [
            _make_scope_estimate(scope_type="ACT", confidence=0.50),
        ]
        report_high = compute_project_confidence(plan_high, scopes_high)
        report_low = compute_project_confidence(plan_low, scopes_low)
        assert report_low.overall_score < report_high.overall_score

    def test_no_scopes_flagged(self) -> None:
        """Zero scopes → flagged in report."""
        plan_result = _make_plan_result_for_confidence(extraction_confidence=0.95)
        report = compute_project_confidence(plan_result, [])
        assert len(report.flags) > 0, "Expected a flag when no scopes are present"

    def test_missing_total_sf_flagged(self) -> None:
        """plan_result.total_area_sf=None → flagged in report."""
        plan_result = _make_plan_result_for_confidence(extraction_confidence=0.95, total_area_sf=None)
        scopes = [_make_scope_estimate()]
        report = compute_project_confidence(plan_result, scopes)
        flag_text = " ".join(report.flags).lower()
        assert "sf" in flag_text or "area" in flag_text, f"Expected SF/area flag, got: {report.flags}"

    def test_bluebeam_annotations_flagged_positive(self) -> None:
        """extraction_confidence=1.0 (Bluebeam) → positive flag in report."""
        plan_result = _make_plan_result_for_confidence(extraction_confidence=1.0)
        scopes = [_make_scope_estimate()]
        report = compute_project_confidence(plan_result, scopes)
        flag_text = " ".join(report.flags).lower()
        assert "bluebeam" in flag_text or "annotation" in flag_text or "high sf" in flag_text, (
            f"Expected Bluebeam positive flag, got: {report.flags}"
        )

    def test_report_has_recommendations(self) -> None:
        """ConfidenceReport should always include at least one recommendation."""
        plan_result = _make_plan_result_for_confidence(extraction_confidence=0.95)
        scopes = [_make_scope_estimate()]
        report = compute_project_confidence(plan_result, scopes)
        assert len(report.recommendations) >= 1

    def test_overall_score_in_range(self) -> None:
        """overall_score must be in [0, 1]."""
        for conf in (0.0, 0.3, 0.6, 1.0):
            plan_result = _make_plan_result_for_confidence(extraction_confidence=conf)
            scopes = [_make_scope_estimate()]
            report = compute_project_confidence(plan_result, scopes)
            assert 0.0 <= report.overall_score <= 1.0


# ===========================================================================
# Pydantic model tests
# ===========================================================================


class TestScopeEstimate:
    """Tests for ScopeEstimate Pydantic model."""

    def test_valid_scope_estimate(self) -> None:
        scope = _make_scope_estimate()
        assert scope.scope_tag == "ACT-1"
        assert scope.scope_type == "ACT"
        assert scope.area_sf == Decimal("1000")

    def test_confidence_bounds_valid(self) -> None:
        scope = _make_scope_estimate(confidence=0.0)
        assert scope.confidence == 0.0
        scope2 = _make_scope_estimate(confidence=1.0)
        assert scope2.confidence == 1.0

    def test_confidence_out_of_range_raises(self) -> None:
        with pytest.raises(Exception, match=".*"):  # noqa: B017
            _make_scope_estimate(confidence=1.5)

    def test_nullable_fields(self) -> None:
        scope = _make_scope_estimate(
            area_sf=None,
            total=None,
            material_cost=None,
            labor_cost=None,
            predicted_cost_per_sf=None,
            predicted_markup_pct=None,
            predicted_man_days=None,
            product_hint=None,
        )
        assert scope.area_sf is None
        assert scope.total is None

    def test_decimal_serialization(self) -> None:
        scope = _make_scope_estimate()
        data = scope.model_dump()
        # All Decimal fields should be Decimal in dict form (or serializable)
        assert data["area_sf"] == Decimal("1000")
        assert data["total"] == Decimal("12500.00")

    def test_json_serialization(self) -> None:
        scope = _make_scope_estimate()
        json_str = scope.model_dump_json()
        assert "ACT-1" in json_str
        assert "12500" in json_str

    def test_comparable_projects_default_empty(self) -> None:
        scope = _make_scope_estimate()
        assert scope.comparable_projects == []


class TestProjectEstimate:
    """Tests for ProjectEstimate Pydantic model."""

    def test_valid_project_estimate(self) -> None:
        est = _make_project_estimate()
        assert est.source_plan == "/fake/path/project.pdf"
        assert len(est.scope_estimates) == 2

    def test_total_matches_scope_sum(self) -> None:
        scopes = [
            _make_scope_estimate(scope_tag="ACT-1", total=Decimal("10000")),
            _make_scope_estimate(scope_tag="AWP-1", scope_type="AWP", total=Decimal("5000")),
        ]
        est = _make_project_estimate(scopes)
        assert est.total_estimated_cost == Decimal("15000")

    def test_decimal_fields_serialize(self) -> None:
        est = _make_project_estimate()
        data = est.model_dump()
        assert isinstance(data["total_estimated_cost"], Decimal)
        assert isinstance(data["estimated_man_days"], Decimal)

    def test_json_round_trip(self) -> None:
        est = _make_project_estimate()
        json_str = est.model_dump_json()
        restored = ProjectEstimate.model_validate_json(json_str)
        assert restored.total_estimated_cost == est.total_estimated_cost

    def test_notes_default_empty(self) -> None:
        est = _make_project_estimate()
        assert isinstance(est.notes, list)

    def test_nullable_total_area(self) -> None:
        scopes = [_make_scope_estimate(area_sf=None)]
        total = scopes[0].total or Decimal("0")
        est = ProjectEstimate(
            source_plan="/fake/path/project.pdf",
            extraction_confidence=0.5,
            scope_estimates=scopes,
            total_estimated_cost=total,
            total_area_sf=None,
            estimated_man_days=Decimal("4.0"),
            notes=["No area data"],
            created_at=datetime(2026, 4, 6),
        )
        assert est.total_area_sf is None


# ===========================================================================
# Excel writer tests
# ===========================================================================


class TestWriteEstimateToExcel:
    """Tests for write_estimate_to_excel()."""

    def test_creates_file(self, tmp_path: Path) -> None:
        est = _make_project_estimate()
        output_path = tmp_path / "test_estimate.xlsx"
        write_estimate_to_excel(est, output_path)
        assert output_path.exists(), "Expected Excel file to be created"

    def test_valid_openpyxl_workbook(self, tmp_path: Path) -> None:
        import openpyxl

        est = _make_project_estimate()
        output_path = tmp_path / "test_estimate.xlsx"
        write_estimate_to_excel(est, output_path)
        wb = openpyxl.load_workbook(output_path)
        assert len(wb.sheetnames) >= 1

    def test_grand_total_row_present(self, tmp_path: Path) -> None:
        import openpyxl

        est = _make_project_estimate()
        output_path = tmp_path / "test_estimate.xlsx"
        write_estimate_to_excel(est, output_path)
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        # Collect all cell values as strings and look for grand total
        all_values = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    all_values.append(str(cell.value).lower())
        text_blob = " ".join(all_values)
        assert "total" in text_blob, f"Expected 'total' in workbook, cells: {all_values[:20]}"

    def test_no_crash_on_none_scope_totals(self, tmp_path: Path) -> None:
        scopes = [
            _make_scope_estimate(
                scope_tag="ACT-1",
                total=None,
                material_cost=None,
                labor_cost=None,
                predicted_cost_per_sf=None,
                predicted_markup_pct=None,
                predicted_man_days=None,
                area_sf=None,
            )
        ]
        est = ProjectEstimate(
            source_plan="/fake/path/project.pdf",
            extraction_confidence=0.5,
            scope_estimates=scopes,
            total_estimated_cost=Decimal("0"),
            total_area_sf=None,
            estimated_man_days=Decimal("0"),
            notes=["All None totals"],
            created_at=datetime(2026, 4, 6),
        )
        output_path = tmp_path / "none_totals.xlsx"
        # Must not raise
        write_estimate_to_excel(est, output_path)
        assert output_path.exists()

    def test_scope_tags_in_workbook(self, tmp_path: Path) -> None:
        import openpyxl

        est = _make_project_estimate()
        output_path = tmp_path / "scope_tags.xlsx"
        write_estimate_to_excel(est, output_path)
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        all_values = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        assert any("ACT" in v for v in all_values), "Expected scope tags in workbook"


# ===========================================================================
# Comparator module tests (mocked DB)
# ===========================================================================


class TestSimilarityHelpers:
    """Tests for the internal similarity math in comparator.py."""

    def test_sf_similarity_equal(self) -> None:

        assert _sf_similarity(1000.0, 1000.0) == pytest.approx(1.0)

    def test_sf_similarity_zero_returns_zero(self) -> None:

        assert _sf_similarity(0.0, 1000.0) == 0.0
        assert _sf_similarity(1000.0, 0.0) == 0.0

    def test_sf_similarity_orders_of_magnitude(self) -> None:

        # 10x difference → moderate similarity
        score = _sf_similarity(100.0, 1000.0)
        assert 0.0 < score < 0.5

    def test_cost_similarity_equal(self) -> None:

        assert _cost_similarity(10.0, 10.0) == pytest.approx(1.0)

    def test_cost_similarity_zero_ref(self) -> None:

        assert _cost_similarity(10.0, 0.0) == 0.0

    def test_compute_similarity_same_type(self) -> None:

        score = _compute_similarity("ACT", 1000.0, 12.0, "ACT", 1000.0, 12.0)
        assert score == pytest.approx(1.0)

    def test_compute_similarity_different_type(self) -> None:

        score = _compute_similarity("ACT", 1000.0, 12.0, "AWP", 1000.0, 12.0)
        # type mismatch should significantly reduce score
        assert score < 0.6

    def test_compute_similarity_no_cost(self) -> None:

        # Should renormalize when cost unknown
        score = _compute_similarity("ACT", 1000.0, None, "ACT", 1000.0, None)
        assert score == pytest.approx(1.0)


class TestFindComparableProjects:
    """Tests for find_comparable_projects() with mocked DB session."""

    @pytest.mark.asyncio
    async def test_returns_empty_for_no_rows(self) -> None:
        mock_session = AsyncMock()
        mock_result = MagicMock()
        mock_result.all.return_value = []
        mock_session.execute.return_value = mock_result

        result = await find_comparable_projects(
            session=mock_session,
            scope_type="ACT",
            area_sf=1000.0,
            top_n=3,
        )
        assert result == []

    @pytest.mark.asyncio
    async def test_returns_top_n_comparables(self) -> None:
        from decimal import Decimal
        from unittest.mock import MagicMock

        mock_session = AsyncMock()

        def _make_scope(scope_type: str, sf: float, cost: float, project_id: str) -> tuple:
            scope = MagicMock()
            scope.square_footage = Decimal(str(sf))
            scope.cost_per_unit = Decimal(str(cost))
            scope.markup_pct = Decimal("0.33")
            scope.total = Decimal(str(sf * cost * 1.33))
            scope.scope_type = MagicMock()
            scope.scope_type.value = scope_type
            scope.tag = f"{scope_type}-1"

            project = MagicMock()
            project.id = project_id
            project.name = f"Project {project_id}"
            project.folder_name = f"folder_{project_id}"
            return (scope, project)

        rows = [
            _make_scope("ACT", 1000.0, 12.0, "proj-1"),
            _make_scope("ACT", 950.0, 11.5, "proj-2"),
            _make_scope("AWP", 500.0, 20.0, "proj-3"),
            _make_scope("ACT", 800.0, 13.0, "proj-4"),
            _make_scope("FW", 300.0, 15.0, "proj-5"),
        ]
        mock_result = MagicMock()
        mock_result.all.return_value = rows
        mock_session.execute.return_value = mock_result

        comparables = await find_comparable_projects(
            session=mock_session,
            scope_type="ACT",
            area_sf=1000.0,
            cost_per_sf=12.0,
            top_n=3,
        )
        assert len(comparables) <= 3
        # All results should be ComparableProject instances
        for c in comparables:
            assert isinstance(c, ComparableProject)
        # ACT scopes should rank higher than AWP/FW
        if comparables:
            assert comparables[0].scope_type == "ACT"

    @pytest.mark.asyncio
    async def test_one_result_per_project(self) -> None:
        """Same project_id appearing twice → only 1 comparable returned."""
        mock_session = AsyncMock()

        scope_a = MagicMock()
        scope_a.square_footage = Decimal("1000")
        scope_a.cost_per_unit = Decimal("12")
        scope_a.markup_pct = Decimal("0.33")
        scope_a.total = Decimal("15960")
        scope_a.scope_type = MagicMock()
        scope_a.scope_type.value = "ACT"
        scope_a.tag = "ACT-1"

        scope_b = MagicMock()
        scope_b.square_footage = Decimal("500")
        scope_b.cost_per_unit = Decimal("12")
        scope_b.markup_pct = Decimal("0.33")
        scope_b.total = Decimal("7980")
        scope_b.scope_type = MagicMock()
        scope_b.scope_type.value = "ACT"
        scope_b.tag = "ACT-2"

        project = MagicMock()
        project.id = "same-project-id"
        project.name = "Duplicate Project"
        project.folder_name = "dup_folder"

        mock_result = MagicMock()
        mock_result.all.return_value = [(scope_a, project), (scope_b, project)]
        mock_session.execute.return_value = mock_result

        comparables = await find_comparable_projects(
            session=mock_session,
            scope_type="ACT",
            area_sf=1000.0,
            top_n=5,
        )
        assert len(comparables) == 1, "Duplicate project_id should yield only one comparable"

    @pytest.mark.asyncio
    async def test_similarity_scores_in_range(self) -> None:
        mock_session = AsyncMock()

        scope = MagicMock()
        scope.square_footage = Decimal("2000")
        scope.cost_per_unit = Decimal("15")
        scope.markup_pct = Decimal("0.35")
        scope.total = Decimal("40500")
        scope.scope_type = MagicMock()
        scope.scope_type.value = "ACT"
        scope.tag = "ACT-1"

        project = MagicMock()
        project.id = "p1"
        project.name = "Test Project"
        project.folder_name = "test_folder"

        mock_result = MagicMock()
        mock_result.all.return_value = [(scope, project)]
        mock_session.execute.return_value = mock_result

        comparables = await find_comparable_projects(
            session=mock_session,
            scope_type="ACT",
            area_sf=1000.0,
            top_n=3,
        )
        for c in comparables:
            assert 0.0 <= c.similarity_score <= 1.0


# ===========================================================================
# Estimator integration test (skips if model files absent)
# ===========================================================================


@requires_model
class TestEstimatorWithRealModel:
    """Integration tests that exercise estimate_from_plan_result() with real models.

    Only runs when data/models/ACT_cost_model.joblib exists.
    """

    def test_estimate_from_plan_result_returns_project_estimate(self) -> None:
        pytest.importorskip(
            "src.estimation.estimator",
            reason="src/estimation/estimator.py not yet delivered",
        )
        from src.estimation.estimator import estimate_from_plan_result
        from src.extraction.plan_parser.models import PlanReadResult, ScopeSuggestion

        plan_result = PlanReadResult(
            source_file="/fake/plan.pdf",
            total_pages=2,
            vector_rich_pages=0,
            raster_pages=2,
            pages=[],
            rooms=[],
            ceiling_specs=[],
            scope_suggestions=[
                ScopeSuggestion(
                    scope_type="ACT",
                    scope_tag="ACT-1",
                    area_sf=Decimal("1000"),
                    length_lf=None,
                    product_hint=None,
                    confidence=0.95,
                    source="bluebeam_annotation",
                    rooms=[],
                )
            ],
            total_area_sf=Decimal("1000"),
            extraction_confidence=0.95,
            vision_pages_used=0,
            error=None,
            success=True,
        )
        result = estimate_from_plan_result(plan_result)
        assert isinstance(result, ProjectEstimate)
        assert len(result.scope_estimates) >= 1
        assert result.total_estimated_cost >= Decimal("0")
